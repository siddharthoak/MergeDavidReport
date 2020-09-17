from openpyxl import load_workbook

date_to_be_updated = "06/30/2021"
sow_id = "3198"
report_file = "/home/siddharthoak/Downloads/CBS_Project_Members_Details_gsLab_15_September_2020.xlsx"
workbook_ibm = load_workbook(filename=report_file)
sheet = workbook_ibm.active
for row in sheet.iter_rows(min_row=2, values_only=False):
    if row[12].value == sow_id:
        row[11].value = date_to_be_updated
        print("Update date for ", row[5].value)

workbook_ibm.save(report_file)
