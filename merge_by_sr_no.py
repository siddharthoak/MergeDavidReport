from openpyxl import load_workbook
from bluepage_user import BluepageUser

bluepages_sheet = "/home/siddharthoak/Downloads/IDExpiryLatestData.xlsx"
target_sheet = "/home/siddharthoak/Downloads/CBS_Project_Members_laptops.xlsx"

workbook_target = load_workbook(filename=target_sheet)
workbook_bluepages = load_workbook(filename=bluepages_sheet)

# hold the name - user dict
srno_to_user = {}

# OPEN expiry dates sheet and load the data in a dict
sheet = workbook_bluepages.active
for row in sheet.iter_rows(min_row=2, values_only=True):

    bp_user = BluepageUser(name=row[0], email=row[4], manager_name=row[3], srno=row[1], lotus=row[5])
    srno_to_user[bp_user.srno] = bp_user

sheet = workbook_target.active

users_not_found = []

for row in sheet.iter_rows():
    srno = row[0].value
    if srno_to_user.__contains__(srno):
        bp_user = srno_to_user[srno]
        #     update the record
        row[2].value = bp_user.email
        row[3].value = bp_user.lotus
    else:
        users_not_found.append(srno)

# Save the changes
workbook_target.save(target_sheet)
