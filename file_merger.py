from openpyxl import load_workbook
from bluepage_user import BluepageUser

bluepages_sheet = "/home/siddharthoak/Downloads/IDExpiryLatestData.xlsx"
david_sheet = "/home/siddharthoak/Downloads/CBS_Project_Members_Details_gsLab_15_September_2020.xlsx"

workbook_david = load_workbook(filename=david_sheet)
workbook_bluepages = load_workbook(filename=bluepages_sheet)

# hold the name - user dict
name_to_user = {}

# OPEN expiry dates sheet and load the data in a dict
sheet = workbook_bluepages.active
for row in sheet.iter_rows(min_row=2, values_only=True):
    bp_user = BluepageUser(name=row[0], srno=row[2], lotus=row[6], email=row[5], manager_name=row[4], expiry_date=row[3])
    name_to_user[bp_user.email] = bp_user

sheet = workbook_david.active

users_not_found = []

for row in sheet.iter_rows(min_row=2, min_col=15):
    name = row[5].value
    email_address = row[7].value
    date_col = row[12].value
    manager_name = row[10].value
    if email_address == None:
        continue
    email_address =email_address.strip()
    print(email_address)
    if (name_to_user.__contains__(email_address)):
        bp_user = name_to_user[email_address]
        #     update the record
        row[12].value = bp_user.expiry_date
        row[11].value = bp_user.manager_name

        print("updated user record : ", bp_user.name, " old date: ", date_col, " to ", bp_user.expiry_date)
    else:
        users_not_found.append(name)

# Save the changes
workbook_david.save(david_sheet)

print("User: ", users_not_found, " not found")
